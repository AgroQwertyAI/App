from sqlite3 import Connection
import json
from datetime import datetime
import base64
import io
import pandas as pd
from collections import defaultdict

def get_pending_messages(setting_id: int, conn: Connection) -> list[dict]:
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM messages_pending WHERE setting_id = ?", 
        (setting_id,)
    )
    pending_messages = cursor.fetchall()
    return pending_messages

def aggregate_messages(pending_messages: list[dict]) -> dict:
    aggregated_data = {}
    for message in pending_messages:
        formatted_data = json.loads(message["formatted_message_text"])
        for key, values in formatted_data.items():
            if key not in aggregated_data:
                aggregated_data[key] = []
            aggregated_data[key].extend(values)

    return aggregated_data

def save_report_to_db(setting_id: int, file: str, conn: Connection) -> int:
    cursor = conn.cursor()
    now = datetime.now().isoformat()

    cursor.execute(
        """
        INSERT INTO reports (setting_id, timedata, file, extra)
        VALUES (?, ?, ?, ?)
        """,
        (setting_id, now, file, "{}")
    )
    report_id = cursor.lastrowid
    return report_id

def get_image_binary_from_base64(base64_string: str) -> tuple[bytes, str]:
    # Parse image type from base64 string if it has proper format
    image_extension = "jpeg"  # Default extension
    if base64_string.startswith('data:image/'):
        mime_type = base64_string.split(';')[0].split('/')[1]
        image_extension = mime_type
        # Clean up the base64 string by removing the prefix
        base64_string = base64_string.split(',')[1]
    
    # Decode base64 image data
    image_binary = base64.b64decode(base64_string)
    return image_binary, image_extension

def save_message_report_to_db(pending_message: dict, report_id: int, conn: Connection) -> bool:
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    cursor.execute(
        """
        INSERT INTO messages_reports (
            sender_phone_number, sender_name, sender_id, 
            setting_id, report_id, timedata, 
            original_message_text, formatted_message_text, images, extra
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            pending_message["sender_phone_number"],
            pending_message["sender_name"],
            pending_message["sender_id"],
            pending_message["setting_id"],
            report_id,
            now,
            pending_message["original_message_text"],
            pending_message["formatted_message_text"],
            pending_message["images"] if pending_message["images"] else '{"images": []}',
            pending_message["extra"] if pending_message["extra"] else "{}"
        )
    )

def delete_pending_messages(setting_id: int, conn: Connection) -> bool:
    cursor = conn.cursor()
    cursor.execute(
        "DELETE FROM messages_pending WHERE setting_id = ?",
        (setting_id,)
    )
    return cursor.rowcount > 0

def convert_dataframe_to_bytes_xlsx(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    
    # Use openpyxl engine for styling
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Sheet1']
        
        # Import required style objects
        from openpyxl.styles import PatternFill
        
        # Create fill styles
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        # Apply green fill to header (first row)
        for cell in worksheet[1]:
            cell.fill = green_fill
        
        # Apply yellow fill to empty cells and format numerical strings
        for row in worksheet.iter_rows(min_row=2):  # Skip header row
            for cell in row:
                if isinstance(cell.value, str):
                    cell_value = cell.value.strip()
                    if cell_value.isdigit():
                        try:
                            num_value = int(cell_value)
                            # Check if number is in millions range
                            if num_value >= 1000000:
                                # Format for millions: maintain thousands separators without suffix
                                cell.number_format = '# ### ### ##0'
                            else:
                                # Regular integer format with space as thousands separator
                                cell.number_format = '# ##0'
                            cell.value = num_value
                        except ValueError:
                            pass
                    elif is_float_string(cell_value):
                        try:
                            num_value = float(cell_value)
                            # Check if float is in millions range
                            if num_value >= 1000000:
                                # Format for millions: maintain thousands separators without suffix
                                cell.number_format = '# ### ### ##0.00'
                            else:
                                # Regular float format with space as thousands separator
                                cell.number_format = '# ##0.00'
                            cell.value = num_value
                        except ValueError:
                            pass
        
        # Make all columns wider
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Add some padding to the maximum length and set the column width
            # Use proportional padding - less padding for longer content
            adjusted_width = min(max_length + 1, max_length * 1.05)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    buffer.seek(0)
    file_content = buffer.read()
    return file_content

def is_float_string(s: str) -> bool:
    """Check if string represents a valid float."""
    try:
        # Will fail if not a valid float format
        float(s)
        # Check if it contains a decimal point
        return '.' in s
    except ValueError:
        return False

def get_aggregated_json_from_messages(messages: list[dict]) -> dict:
    aggregated_json = {}
    for message in messages:
        message_json = json.loads(message["formatted_message_text"])
        for field, values in message_json.items():
            if field not in aggregated_json:
                aggregated_json[field] = []
            aggregated_json[field].extend(values)

    return aggregated_json

def group_messages_by_sender(messages: list[dict]) -> dict:
    # Group messages by sender_id
    messages_by_sender = defaultdict(list)
    for message in messages:
        sender_id = message["sender_id"]
        messages_by_sender[sender_id].append(message)

    return messages_by_sender